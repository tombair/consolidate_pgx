#!/usr/bin/python

import openpyxl
import csv
import sys
import re

wb = openpyxl.load_workbook(filename=sys.argv[1])

sheets = wb.get_sheet_names()

phenodata = {}
meds = {}
def mednametruc(name):
    mapping = {'.':'', ',':'','\t':'','\n':'',' ':'', 'mg':'','tablet':'','oral':'','MG':'','patch':'', 'ML':'', 'mL':'', 'powder':'','capsule':'', 'IV':'', 'drops':''}
    for k,v in mapping.iteritems():
        name = name.replace(k,v)
    name = re.sub('\d', '', name)
    return name

    
drug_list = [] #a list of all possible drugs prepend Adverse_ to the front of drugs used in response to adverse events
ws =wb.get_sheet_by_name('Med Admin for Adverse Visits')
for r in ws.iter_rows():
    medName = 'Adverse_'+r[8].value
    medName = mednametruc(medName)
    drug_list.append(medName)
ws = wb.get_sheet_by_name('All Rx Medications')
for r in ws.iter_rows():
    medName = r[4].value
    if medName is not None:
        medName = mednametruc(medName)
        drug_list.append(medName)
ws = wb.get_sheet_by_name('All Pt Reported Medications')
for r in ws.iter_rows():
    medName = r[4].value
    if medName is not None:
        medName = mednametruc(medName)
        drug_list.append(medName)
drug_list = list(set(drug_list))
meds['ID'] = drug_list

ws =wb.get_sheet_by_name('Med Admin for Adverse Visits')
for r in ws.iter_rows():
    ptId = r[0].value
    medName = 'Adverse_'+r[8].value
    medName = mednametruc(medName)
    if not meds.has_key(ptId):
        meds[ptId] = [0]*len(drug_list)
    index_pos = drug_list.index(medName)
    meds[ptId][index_pos] = 1
ws = wb.get_sheet_by_name('All Rx Medications')
for r in ws.iter_rows():
    ptId = r[0].value
    medName = r[4].value
    if medName is not None:
        medName = mednametruc(medName)
        if not meds.has_key(ptId):
            meds[ptId] = [0]*len(drug_list)
        index_pos = drug_list.index(medName)
        meds[ptId][index_pos] = 1
ws = wb.get_sheet_by_name('All Pt Reported Medications')
for r in ws.iter_rows():
    ptId = r[0].value
    medName = r[4].value
    if medName is not None:
        medName = mednametruc(medName)
        if not meds.has_key(ptId):
            meds[ptId] = [0]*len(drug_list)
        index_pos = drug_list.index(medName)
        meds[ptId][index_pos] = 1
#now should have a drug use matrix indexed by ptID full of zeros (mostly) and ones

ws = wb.get_sheet_by_name('Patient Visits')
#go through and get pt id and total up the ammount spent (I) and LOS (E)
for r in ws.iter_rows():
    if not phenodata.has_key(r[0].value):
        phenodata[r[0].value] = {"Total_Charges":'NA',"LOS":'NA',"DX_Name":'NA',"Adverse_ICD9":'NA'}
    phenodata[r[0].value]['Total_Charges'] = r[8].value
    phenodata[r[0].value]['LOS'] = r[4].value

ws = wb.get_sheet_by_name('DX')
# go through and associate ptid with DX_NAME(D) and E and F and G
for r in ws.iter_rows():
    if phenodata.has_key(r[0].value):
        phenodata[r[0].value]['DX_Name'] = r[3].value.replace(',',' ')
        phenodata[r[0].value]['Adverse_ICD9'] = r[6].value.replace(',',' ')

#need to make sure phenodata only has ptids that are in meds
newpheno = {}
for x in phenodata:
    if meds.has_key(x):
        newpheno[x] = phenodata[x]
with open("phenodata.csv", 'w') as pheno_out:
    header = ["Total_Charges","LOS","DX_Name","Adverse_ICD9"]
    pheno_out.write("ID,"+",".join(header))
    pheno_out.write("\n")
    for x in newpheno:
        out = [x]
        for h in header:
            try:
                out.append(newpheno[x][h])
            except:
                pass
        pheno_out.write(",".join(map(str,out)))
        pheno_out.write("\n")
with open("drug_data.csv",'w') as drug_out:
    drug_out.write("ID\t"+"\t".join(map(str,meds['ID'])))
    drug_out.write("\n")
    for id in newpheno:
        try:
            drug_out.write(id+"\t"+"\t".join(map(str,meds[id])))
            drug_out.write("\n")
        except:
            pass #in meds but not newpheno
        




        
    

    
